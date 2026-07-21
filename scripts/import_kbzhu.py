"""
Разовый импорт карточек КБЖУ (Сценарий Г) из Excel-таблицы в inventory.db.

Использование:
    python scripts/import_kbzhu.py <путь_к_файлу.xlsx>              # dry-run (ничего не пишет в БД)
    python scripts/import_kbzhu.py <путь_к_файлу.xlsx> --commit     # реальная запись

Формат исходного листа (КБЖУ_food.xlsx):
    Заголовок — 2 строки, данные с 3-й строки.
    A: название блюда
    B: выход блюда, кг. (не импортируется — нет такого поля в product_info)
    C: жиры, г. (на 1 порцию)
    D: белки, г. (на 1 порцию)
    E: углеводы, г. (на 1 порцию)
    F: калорийность, ккал. (на 1 порцию)

drinks.xlsx пока сознательно не импортируется (другая структура — варианты по
типу молока в отдельном столбце), это отдельная задача.
"""
import argparse
import os
import sqlite3
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import DB_PATH  # noqa: E402

_SCHEMA_PATH = os.path.join(os.path.dirname(__file__), "..", "database", "schema.sql")

NAME_OVERRIDES: dict[str, str] = {
    "Круассан класический": "Круассан классический",
}


class KbzhuRow:
    def __init__(self, raw_name: str, fat, protein, carbs, calories):
        self.raw_name = raw_name
        self.name = NAME_OVERRIDES.get(raw_name.strip(), raw_name.strip())
        self.normalized_name = self.name.lower()
        self.fat = fat
        self.protein = protein
        self.carbs = carbs
        self.calories = calories
        self.skipped_reason: str | None = None


def _to_number(value) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    return None


def read_rows(xlsx_path: str) -> list[KbzhuRow]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows: list[KbzhuRow] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        name, _vykhod, fat, protein, carbs, calories = row
        if name is None:
            continue
        rows.append(
            KbzhuRow(
                raw_name=name,
                fat=_to_number(fat),
                protein=_to_number(protein),
                carbs=_to_number(carbs),
                calories=_to_number(calories),
            )
        )

    return rows


def process_rows(rows: list[KbzhuRow]) -> list[KbzhuRow]:
    """Дедупликация по нормализованному названию."""
    seen_names: dict[str, KbzhuRow] = {}
    result: list[KbzhuRow] = []

    for row in rows:
        if row.normalized_name in seen_names:
            row.skipped_reason = f"дубликат названия (уже есть строка '{seen_names[row.normalized_name].raw_name}')"
            result.append(row)
            continue

        seen_names[row.normalized_name] = row
        result.append(row)

    return result


def print_summary(rows: list[KbzhuRow]) -> None:
    ok_rows = [r for r in rows if r.skipped_reason is None]
    skipped_rows = [r for r in rows if r.skipped_reason is not None]

    print(f"Всего строк прочитано: {len(rows)}")
    print(f"К импорту: {len(ok_rows)}")
    print(f"Пропущено: {len(skipped_rows)}")
    print()

    if skipped_rows:
        print("=== Пропущенные строки ===")
        for r in skipped_rows:
            print(f"  {r.raw_name!r}: {r.skipped_reason}")
        print()

    print("=== Все карточки к импорту ===")
    for r in ok_rows:
        print(
            f"  {r.name!r}: {r.calories} ккал "
            f"(Б: {r.protein}г, Ж: {r.fat}г, У: {r.carbs}г)"
        )


def commit_to_db(rows: list[KbzhuRow]) -> None:
    ok_rows = [r for r in rows if r.skipped_reason is None]

    with open(_SCHEMA_PATH, "r", encoding="utf-8") as f:
        schema = f.read()

    conn = sqlite3.connect(DB_PATH)
    try:
        conn.executescript(schema)
        conn.execute("DELETE FROM product_info")

        inserted = 0
        for r in ok_rows:
            conn.execute(
                "INSERT INTO product_info (name, calories, protein, fat, carbs) VALUES (?, ?, ?, ?, ?)",
                (r.name, r.calories, r.protein, r.fat, r.carbs),
            )
            inserted += 1

        conn.commit()
        print(f"Записано/обновлено карточек: {inserted}")
    finally:
        conn.close()


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx_path", help="Путь к Excel-файлу с КБЖУ (food)")
    parser.add_argument(
        "--commit", action="store_true", help="Реально записать данные в inventory.db"
    )
    args = parser.parse_args()

    rows = read_rows(args.xlsx_path)
    rows = process_rows(rows)
    print_summary(rows)

    if args.commit:
        print()
        commit_to_db(rows)
    else:
        print()
        print("Dry-run завершён, в БД ничего не записано. Добавьте --commit для реального импорта.")


if __name__ == "__main__":
    main()

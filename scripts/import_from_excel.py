"""
Разовый импорт товаров/остатков из Excel-таблицы в inventory.db.

Использование:
    python scripts/import_from_excel.py <путь_к_файлу.xlsx>              # dry-run (ничего не пишет в БД)
    python scripts/import_from_excel.py <путь_к_файлу.xlsx> --commit     # реальная запись

Формат исходного листа (см. TECH_SPEC.md):
    A: категория (заполнена только в первой строке своей группы)
    B: название товара
    C: поставщик (игнорируется)
    D: остаток Кофейня
    E: единица измерения Кофейня
    F: остаток Склад
    G: единица измерения Склад
"""
import argparse
import os
import sqlite3
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import DB_PATH  # noqa: E402

_SCHEMA_PATH = os.path.join(os.path.dirname(__file__), "..", "database", "schema.sql")

EXCLUDED_NAMES = {"соевое"}

EXCLUDED_CATEGORIES = {"остатки"}

UNIT_OVERRIDES: dict[str, str] = {
    "Малина с/м": "уп",
    "Nika универсал": "рулон",
    "Таблетки для кофемолок": "шт",
    "контейнер для чизкейков": "шт",
    "соуснички": "шт",
    "Влажные салфетки": "шт",
    "Пакет SNA": "шт",
    "Сахар в стиках": "шт",
    "Лимонад Матча эликсир": "шт",
    "кола": "шт",
    "брауни cbd черная соль": "шт",
    "Shakabar Соленая карамель": "шт",
    "Shakabar черничный пай": "шт",
    "Shakabar банановый кекс": "шт",
    "Shakabar шоколадное печенье": "шт",
    "Shakabar миндаль клюква": "шт",
    "шоколад перу 70": "шт",
    "шоколад raspberry": "шт",
    "шоколад кокос ананас": "шт",
    "Шоколад дабл кор": "шт",
    "Шоколад софт блум": "шт",
    "сэндвич с фисташковым суфле": "шт",
    "Дрипы Африка": "уп",
    "мексика сочиапа": "уп",
    "Бразилия Педра вермеля": "уп",
    "Наклейки сырники": "уп",
    "лимонад ананас": "шт",
    "Матча элексир А/Ч": "шт",
    "Матча элексир К/И": "шт",
    "Матче элексир": "шт",
    "Трюфель Кешью": "шт",
    "Трюфель Грецкий": "шт",
    "Ириска": "шт",
    "Калитка абрикос": "шт",
    "Калитка вишня": "шт",
    "вупи матча": "шт",
    "вупи ск": "шт",
    "Драфт slowBrew": "шт",
}

CATEGORY_NAME_PREFIX = {
    "молоко": "Молоко",
}


class ImportRow:
    def __init__(self, category: str, raw_name: str, cafe_qty, cafe_unit, wh_qty, wh_unit):
        self.category = category
        self.raw_name = raw_name
        self.cafe_qty = cafe_qty
        self.cafe_unit = cafe_unit
        self.wh_qty = wh_qty
        self.wh_unit = wh_unit
        self.skipped_reason: str | None = None
        self.display_name = raw_name
        self.normalized_name = raw_name.strip().lower()
        self.unit: str | None = None


def _to_quantity(value) -> float:
    """Числовой остаток -> float. Нечисловое/пустое значение -> 0.0 (данные не заполнены)."""
    if isinstance(value, (int, float)):
        return float(value)
    return 0.0


def read_rows(xlsx_path: str) -> list[ImportRow]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows: list[ImportRow] = []
    current_category: str | None = None

    for row in ws.iter_rows(min_row=3, values_only=True):
        a, b, _c, d, e, f, g = row
        if a:
            current_category = a.strip()
        if b is None:
            continue

        raw_name = b.strip()
        rows.append(
            ImportRow(
                category=current_category,
                raw_name=raw_name,
                cafe_qty=d,
                cafe_unit=e,
                wh_qty=f,
                wh_unit=g,
            )
        )

    return rows


def process_rows(rows: list[ImportRow]) -> list[ImportRow]:
    """Применяет исключения, override единиц, склейку display_name, дедупликацию."""
    seen_names: dict[str, ImportRow] = {}
    result: list[ImportRow] = []

    for row in rows:
        if row.normalized_name in EXCLUDED_NAMES:
            row.skipped_reason = "исключён из импорта (снят с реализации)"
            result.append(row)
            continue

        if row.category in EXCLUDED_CATEGORIES:
            row.skipped_reason = f"категория {row.category!r} исключена из импорта (снята с реализации)"
            result.append(row)
            continue

        prefix = CATEGORY_NAME_PREFIX.get(row.category)
        if prefix and not row.raw_name.lower().startswith(prefix.lower()):
            lowered = row.raw_name[0].lower() + row.raw_name[1:]
            row.display_name = f"{prefix} {lowered}"
        row.normalized_name = row.display_name.strip().lower()

        if row.raw_name in UNIT_OVERRIDES:
            row.unit = UNIT_OVERRIDES[row.raw_name]
        elif row.cafe_unit and row.wh_unit and row.cafe_unit != row.wh_unit:
            row.unit = row.wh_unit
        else:
            row.unit = row.wh_unit or row.cafe_unit

        if row.normalized_name in seen_names:
            row.skipped_reason = f"дубликат названия (уже есть строка '{seen_names[row.normalized_name].raw_name}')"
            result.append(row)
            continue

        seen_names[row.normalized_name] = row
        result.append(row)

    return result


def print_summary(rows: list[ImportRow]) -> None:
    ok_rows = [r for r in rows if r.skipped_reason is None]
    skipped_rows = [r for r in rows if r.skipped_reason is not None]
    categories = sorted({r.category for r in ok_rows if r.category})

    print(f"Всего строк прочитано: {len(rows)}")
    print(f"К импорту: {len(ok_rows)}")
    print(f"Пропущено: {len(skipped_rows)}")
    print(f"Категорий: {len(categories)} -> {categories}")
    print()

    if skipped_rows:
        print("=== Пропущенные строки ===")
        for r in skipped_rows:
            print(f"  {r.raw_name!r}: {r.skipped_reason}")
        print()

    print("=== Примеры товаров (первые 15) ===")
    for r in ok_rows[:15]:
        print(
            f"  [{r.category}] {r.raw_name!r} -> display_name={r.display_name!r}, "
            f"unit={r.unit!r}, кофейня={_to_quantity(r.cafe_qty)}, склад={_to_quantity(r.wh_qty)}"
        )


def commit_to_db(rows: list[ImportRow]) -> None:
    ok_rows = [r for r in rows if r.skipped_reason is None]

    with open(_SCHEMA_PATH, "r", encoding="utf-8") as f:
        schema = f.read()

    conn = sqlite3.connect(DB_PATH)
    try:
        conn.executescript(schema)

        category_ids: dict[str, int] = {}
        for r in ok_rows:
            if not r.category or r.category in category_ids:
                continue
            conn.execute(
                "INSERT OR IGNORE INTO categories (name) VALUES (?)", (r.category,)
            )
            cur = conn.execute(
                "SELECT id FROM categories WHERE name = ?", (r.category,)
            )
            category_ids[r.category] = cur.fetchone()[0]

        inserted = 0
        for r in ok_rows:
            category_id = category_ids.get(r.category)
            conn.execute(
                "INSERT INTO products (name, display_name, unit, category_id) VALUES (?, ?, ?, ?)",
                (r.normalized_name, r.display_name, r.unit, category_id),
            )
            product_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

            conn.execute(
                "INSERT INTO stock (product_id, location, quantity) VALUES (?, 'coffee_shop', ?)",
                (product_id, _to_quantity(r.cafe_qty)),
            )
            conn.execute(
                "INSERT INTO stock (product_id, location, quantity) VALUES (?, 'warehouse', ?)",
                (product_id, _to_quantity(r.wh_qty)),
            )
            inserted += 1

        conn.commit()
        print(f"Записано товаров: {inserted}")
    finally:
        conn.close()


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx_path", help="Путь к Excel-файлу с товарами")
    parser.add_argument(
        "--commit", action="store_true", help="Реально записать данные в inventory.db"
    )
    args = parser.parse_args()

    rows = read_rows(args.xlsx_path)
    rows = process_rows(rows)
    print_summary(rows)

    if args.commit:
        print()
        commit_to_db(rows)
    else:
        print()
        print("Dry-run завершён, в БД ничего не записано. Добавьте --commit для реального импорта.")


if __name__ == "__main__":
    main()
